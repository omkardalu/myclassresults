import requests
import pandas as pd
import re
import time
from bs4 import BeautifulSoup
from io import BytesIO
import PyPDF2
from urllib.parse import urljoin
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from queue import Queue
import socket
import ssl
import urllib3
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configuration
BASE_URL = "https://sbtet.ap.gov.in/APSBTET/results.do"
MAX_CONCURRENT_REQUESTS = 2
FALLBACK_URL = "http://sbtet.ap.gov.in/APSBTET/results.do"
REQUEST_TIMEOUT = 30  # Increased timeout
MAX_RETRIES = 3

class OptimizedSBTETScraper:
    def __init__(self):
        """Initialize the optimized scraper."""
        self.session = self._create_session()
        self.form_data = None
        self._progress = 0
        self._total_pins = 0
        self._lock = threading.Lock()
        self.working_url = None
        
    def _create_session(self):
        """Create a robust session with retry strategy."""
        session = requests.Session()
        
        # Configure retry strategy
        retry_strategy = Retry(
            total=MAX_RETRIES,
            backoff_factor=2,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "POST"]
        )
        
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        
        # Comprehensive headers
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Cache-Control': 'max-age=0',
        })
        
        return session
    
    def get_progress(self):
        """Thread-safe progress getter."""
        with self._lock:
            return (self._progress / self._total_pins * 100) if self._total_pins > 0 else 0
    
    def _update_progress(self):
        """Thread-safe progress updater."""
        with self._lock:
            self._progress += 1
    
    def test_connectivity(self):
        """Test connectivity to SBTET website with multiple approaches."""
        urls_to_test = [BASE_URL, FALLBACK_URL]
        
        for url in urls_to_test:
            try:
                logger.info(f"Testing connectivity to: {url}")
                
                # Test with different timeouts
                for timeout in [10, 20, 30]:
                    try:
                        response = self.session.get(url, timeout=timeout)
                        if response.status_code == 200:
                            logger.info(f"✓ Connected to {url} (timeout: {timeout}s)")
                            self.working_url = url
                            return True, url
                    except requests.exceptions.Timeout:
                        logger.warning(f"Timeout ({timeout}s) for {url}")
                        continue
                    except Exception as e:
                        logger.warning(f"Error with {url}: {e}")
                        break
                        
            except Exception as e:
                logger.error(f"Failed to connect to {url}: {e}")
                continue
                
        return False, None

    def diagnose_connection_issue(self):
        """Diagnose connection issues with detailed information."""
        diagnosis = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "tests": {}
        }
        
        # Test DNS resolution
        try:
            hostname = "sbtet.ap.gov.in"
            ip = socket.gethostbyname(hostname)
            diagnosis["tests"]["dns_resolution"] = {"status": "success", "ip": ip}
        except Exception as e:
            diagnosis["tests"]["dns_resolution"] = {"status": "failed", "error": str(e)}
        
        # Test port connectivity
        for port, protocol in [(80, "HTTP"), (443, "HTTPS")]:
            try:
                sock = socket.create_connection(("sbtet.ap.gov.in", port), timeout=10)
                sock.close()
                diagnosis["tests"][f"{protocol.lower()}_port"] = {"status": "success", "port": port}
            except Exception as e:
                diagnosis["tests"][f"{protocol.lower()}_port"] = {"status": "failed", "port": port, "error": str(e)}
        
        # Test HTTP requests
        for url in [BASE_URL, FALLBACK_URL]:
            try:
                response = self.session.get(url, timeout=15)
                diagnosis["tests"][f"http_request_{url.split('://')[0]}"] = {
                    "status": "success", 
                    "url": url, 
                    "status_code": response.status_code,
                    "response_size": len(response.content)
                }
            except Exception as e:
                diagnosis["tests"][f"http_request_{url.split('://')[0]}"] = {
                    "status": "failed", 
                    "url": url, 
                    "error": str(e)
                }
        
        return diagnosis


    def analyze_form_structure(self):
        """Enhanced form analysis with fallback URLs and better error handling."""
        if self.form_data:
            return self.form_data
        
        # Test connectivity first
        if not self.working_url:
            is_connected, working_url = self.test_connectivity()
            if not is_connected:
                raise Exception("Cannot connect to SBTET website. Check network connectivity.")
            self.working_url = working_url
        
        try:
            logger.info(f"Analyzing form structure from: {self.working_url}")
            
            # Try multiple approaches
            approaches = [
                {"timeout": 30, "verify": True},
                {"timeout": 45, "verify": True},
                {"timeout": 30, "verify": False},  # Disable SSL verification as last resort
            ]
            
            for approach in approaches:
                try:
                    response = self.session.get(
                        self.working_url, 
                        timeout=approach["timeout"],
                        verify=approach["verify"]
                    )
                    response.raise_for_status()
                    
                    # Check if we got actual HTML content
                    if len(response.content) < 100:
                        raise Exception("Response too short - likely not the form page")
                    
                    soup = BeautifulSoup(response.content, 'html.parser')
                    form = soup.find('form')
                    
                    if not form:
                        # Try to find any form-like structures
                        forms = soup.find_all(['form', 'div'], {'class': re.compile(r'form', re.I)})
                        if forms:
                            form = forms[0]
                        else:
                            raise Exception("No form found on the page")
                    
                    # Extract form data
                    self.form_data = {
                        'action': form.get('action', ''),
                        'method': form.get('method', 'POST'),
                        'hidden_fields': {},
                        'base_url': self.working_url
                    }
                    
                    # Store hidden fields
                    for inp in form.find_all('input', type='hidden'):
                        if inp.get('name'):
                            self.form_data['hidden_fields'][inp.get('name')] = inp.get('value', '')
                    
                    # Store other form inputs
                    for inp in form.find_all('input'):
                        input_type = inp.get('type', '').lower()
                        if input_type in ['text', 'number', 'email']:
                            name = inp.get('name', '')
                            if name and name not in self.form_data['hidden_fields']:
                                self.form_data['hidden_fields'][name] = inp.get('value', '')
                    
                    logger.info(f"✓ Form analysis successful! Found {len(self.form_data['hidden_fields'])} fields")
                    return self.form_data
                    
                except requests.exceptions.SSLError as e:
                    logger.warning(f"SSL Error: {e}")
                    if approach["verify"]:
                        continue  # Try next approach
                    else:
                        raise
                except requests.exceptions.Timeout as e:
                    logger.warning(f"Timeout ({approach['timeout']}s): {e}")
                    continue
                except Exception as e:
                    logger.warning(f"Approach failed: {e}")
                    continue
            
            raise Exception("All form analysis approaches failed")
            
        except Exception as e:
            logger.error(f"Form analysis failed: {e}")
            # Provide detailed error information
            if "timeout" in str(e).lower():
                raise Exception(f"Connection timeout to SBTET website. The server may be slow or your network connection is limited. Error: {e}")
            elif "ssl" in str(e).lower():
                raise Exception(f"SSL/Certificate error. Try accessing the website directly to check if it's working. Error: {e}")
            elif "network" in str(e).lower() or "unreachable" in str(e).lower():
                raise Exception(f"Network connectivity issue. Your deployment environment cannot reach the SBTET website. Error: {e}")
            else:
                raise Exception(f"Form analysis failed: {e}")
    
    def generate_pins(self, year, branch_code, college_code, start=1, end=67):
        """Generate PIN list efficiently."""
        return [f"{year}{college_code}-{branch_code}-{str(i).zfill(3)}" for i in range(start, end)]
    
    def _submit_single_request(self, pin, semester="5"):
        """Submit single optimized request."""
        try:
            if not self.form_data:
                if not self.analyze_form_structure():
                    return None
            
            # Prepare minimal POST data
            post_data = self.form_data['hidden_fields'].copy()
            post_data.update({
                'mode': 'getData',
                'aadhar1': pin,
                'grade2': semester
            })
            
            form_url = self.form_data['action']
            if not form_url.startswith('http'):
                form_url = urljoin(BASE_URL, form_url) if form_url else BASE_URL
            
            response = self.session.post(
                form_url,
                data=post_data,
                timeout=REQUEST_TIMEOUT,
                allow_redirects=True
            )
            response.raise_for_status()
            
            content_type = response.headers.get('content-type', '').lower()
            
            if 'application/pdf' in content_type:
                return self._parse_pdf_response(pin, BytesIO(response.content))
            elif 'text/html' in content_type:
                return self._parse_html_response(pin, response.content)
            
            return None
                
        except Exception as e:
            logger.warning(f"Request failed for PIN {pin}: {e}")
            return None
    
    def _parse_pdf_response(self, pin, pdf_content):
        """Optimized PDF parsing."""
        try:
            reader = PyPDF2.PdfReader(pdf_content)
            text = "".join(page.extract_text() or "" for page in reader.pages)
            
            # Extract name efficiently
            name_match = re.search(r"Name\s+([A-Z\s]+)", text)
            name = name_match.group(1).strip() if name_match else "Unknown"
            
            # Extract subjects and marks in one pass
            result = {'PIN': pin, 'NAME': name}
            subjects = []
            
            for line in text.split("\n"):
                line = line.strip()
                match = re.match(r"(\d{3})\s+(\d+)\s+(\d+)([PF])", line)
                if match:
                    subject_code = match.group(1)
                    if subject_code not in subjects:
                        subjects.append(subject_code)
                    
                    ext = int(match.group(2))
                    combined = match.group(3)
                    res = match.group(4)
                    
                    # Parse internal and total marks
                    int_marks, tot = self._parse_combined_marks(combined, ext)
                    
                    result.update({
                        f"{subject_code}_EXT": ext,
                        f"{subject_code}_INT": int_marks,
                        f"{subject_code}_TOT": tot,
                        f"{subject_code}_RES": res
                    })
            
            # Extract totals
            total_match = re.search(r"GrandTotal\s+(\d+)", text)
            result_match = re.search(r"Result\s+(PASS|FAIL)", text)
            
            result.update({
                'TOTAL': int(total_match.group(1)) if total_match else 0,
                'OVERALL_RESULT': 'P' if result_match and result_match.group(1) == 'PASS' else 'F',
                'subjects': subjects
            })
            
            return result
            
        except Exception as e:
            logger.error(f"PDF parsing failed for {pin}: {e}")
            return None
    
    def _parse_html_response(self, pin, html_content):
        """Optimized HTML parsing."""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Quick error check
            if soup.find(text=re.compile(r"error|invalid|not found", re.I)):
                return None
            
            # Check for PDF link
            pdf_link = soup.find('a', href=re.compile(r'\.pdf', re.I))
            if pdf_link:
                pdf_url = pdf_link['href']
                if not pdf_url.startswith('http'):
                    pdf_url = urljoin(BASE_URL, pdf_url)
                
                pdf_response = self.session.get(pdf_url, timeout=REQUEST_TIMEOUT)
                if pdf_response.status_code == 200:
                    return self._parse_pdf_response(pin, BytesIO(pdf_response.content))
            
            # Parse HTML table data
            result = {'PIN': pin, 'NAME': 'Unknown', 'subjects': []}
            
            # Extract name
            text_content = soup.get_text()
            name_match = re.search(r"Name\s*:\s*([A-Z\s]+)", text_content, re.I)
            if name_match:
                result['NAME'] = name_match.group(1).strip()
            
            # Extract subjects and marks from tables
            subjects = []
            for table in soup.find_all('table'):
                for row in table.find_all('tr'):
                    cells = [cell.get_text().strip() for cell in row.find_all(['td', 'th'])]
                    if len(cells) >= 4:
                        subject_match = re.match(r'(\d{3})', cells[0])
                        if subject_match:
                            subject = subject_match.group(1)
                            if subject not in subjects:
                                subjects.append(subject)
                            
                            try:
                                ext = int(re.findall(r'\d+', cells[1])[0]) if re.findall(r'\d+', cells[1]) else 0
                                int_marks = int(re.findall(r'\d+', cells[2])[0]) if re.findall(r'\d+', cells[2]) else 0
                                total = int(re.findall(r'\d+', cells[3])[0]) if re.findall(r'\d+', cells[3]) else 0
                                res = cells[4].upper() if len(cells) > 4 else 'F'
                                
                                result.update({
                                    f"{subject}_EXT": ext,
                                    f"{subject}_INT": int_marks,
                                    f"{subject}_TOT": total,
                                    f"{subject}_RES": res
                                })
                            except (ValueError, IndexError):
                                continue
            
            # Extract totals
            total_match = re.search(r"Total\s*:?\s*(\d+)", text_content, re.I)
            result_match = re.search(r"Result\s*:?\s*(PASS|FAIL)", text_content, re.I)
            
            result.update({
                'TOTAL': int(total_match.group(1)) if total_match else 0,
                'OVERALL_RESULT': 'P' if result_match and result_match.group(1).upper() == 'PASS' else 'F',
                'subjects': subjects
            })
            
            return result
            
        except Exception as e:
            logger.error(f"HTML parsing failed for {pin}: {e}")
            return None
    
    def _parse_combined_marks(self, combined, ext):
        """Helper to parse combined internal+total marks."""
        if len(combined) >= 3:
            for split_pos in [2, 1]:
                if split_pos < len(combined):
                    int_candidate = int(combined[:split_pos])
                    tot_candidate = int(combined[split_pos:])
                    if tot_candidate == ext + int_candidate:
                        return int_candidate, tot_candidate
        
        # Fallback
        return (int(combined[:-2]) if len(combined) > 2 else 0, 
                int(combined[-2:]) if len(combined) >= 2 else int(combined))
    
    def _process_pin_batch(self, pins, semester):
        """Process a batch of PINs with controlled concurrency."""
        results = []
        
        def process_single_pin(pin):
            try:
                result = self._submit_single_request(pin, semester)
                self._update_progress()
                return result
            except Exception as e:
                logger.error(f"Error processing PIN {pin}: {e}")
                self._update_progress()
                return None
        
        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_REQUESTS) as executor:
            future_to_pin = {executor.submit(process_single_pin, pin): pin for pin in pins}
            
            for future in as_completed(future_to_pin):
                pin = future_to_pin[future]
                try:
                    result = future.result(timeout=30)
                    if result:
                        results.append(result)
                        logger.info(f"✓ Success: {pin}")
                    else:
                        logger.warning(f"✗ Failed: {pin}")
                except Exception as e:
                    logger.error(f"Exception for PIN {pin}: {e}")
        
        return results
    
    def scrape_results(self, year, branch_code, college_code, pin_range=(1, 67), semester="5"):
        """Main optimized scraping method."""
        pins = self.generate_pins(year, branch_code, college_code, pin_range[0], pin_range[1])
        self._total_pins = len(pins)
        self._progress = 0
        
        logger.info(f"Starting optimized scraping for {len(pins)} PINs...")
        
        # Analyze form once
        if not self.analyze_form_structure():
            raise Exception("Failed to analyze form structure")
        
        # Process in batches to manage memory
        batch_size = 10
        all_results = []
        
        for i in range(0, len(pins), batch_size):
            batch = pins[i:i + batch_size]
            logger.info(f"Processing batch {i//batch_size + 1}/{(len(pins)-1)//batch_size + 1}")
            
            batch_results = self._process_pin_batch(batch, semester)
            all_results.extend(batch_results)
            
            # Small delay between batches to be respectful
            if i + batch_size < len(pins):
                time.sleep(0.5)
        
        logger.info(f"Scraping complete! Successfully processed: {len(all_results)}/{len(pins)}")
        
        if all_results:
            return self._create_excel_file(all_results)
        else:
            raise Exception("No results obtained")
    
    def _create_excel_file(self, results):
        """Create optimized Excel file.""" 
        if not results:
            return None
        
        # Collect all subjects
        all_subjects = sorted(set().union(*(r.get('subjects', []) for r in results)))
        
        # Build columns
        columns = ['SI.NO', 'PINNUMBERS', 'NAME']
        for subject in all_subjects:
            columns.extend([f"{subject}_EXT", f"{subject}_INT", f"{subject}_TOT", f"{subject}_RES"])
        columns.extend(['TOTAL', 'OVERALL_RESULT'])
        
        # Prepare data efficiently
        data = []
        for i, result in enumerate(results, 1):
            row = [i, result['PIN'], result['NAME']]
            
            for subject in all_subjects:
                row.extend([
                    result.get(f"{subject}_EXT", 0),
                    result.get(f"{subject}_INT", 0),
                    result.get(f"{subject}_TOT", 0),
                    result.get(f"{subject}_RES", 'AB')
                ])
            
            row.extend([result.get('TOTAL', 0), result.get('OVERALL_RESULT', 'F')])
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=columns)
        
        # Save to BytesIO for memory efficiency
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False, engine='openpyxl')
        
        # Add subject headers and formatting
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Insert subject header row
        ws.insert_rows(1)
        col_idx = 4
        
        for subject in all_subjects:
            ws.cell(row=1, column=col_idx, value=subject)
            # Merge cells for subject header
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 3)
            ws.merge_cells(f"{start_col}1:{end_col}1")
            col_idx += 4
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Save to buffer
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)
        
        logger.info(f"Excel file created with {len(results)} students")
        return final_buffer

# Factory function for easy instantiation
def create_scraper():
    """Factory function to create scraper instance."""
    return OptimizedSBTETScraper()

# Example usage function
def scrape_and_download(year='22', college_code='008', branch_code='CM', 
                       pin_range=(1, 70), semester="5"):
    """
    Convenient function to scrape results and return Excel file buffer.
    
    Args:
        year: Year code (e.g., '22')
        college_code: College code (e.g., '008') 
        branch_code: Branch code (e.g., 'CM')
        pin_range: Tuple of (start, end) PIN numbers
        semester: Semester number
    
    Returns:
        BytesIO: Excel file buffer ready for download
    """
    scraper = create_scraper()
    return scraper.scrape_results(year, branch_code, college_code, pin_range, semester)